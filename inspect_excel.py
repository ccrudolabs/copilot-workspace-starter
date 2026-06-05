import openpyxl
import os

path = 'GitHub_Copilot_SKUs_Licenciamiento_360_Carlos_Crudo_v7.xlsx'
print('exists', os.path.exists(path), 'size', os.path.getsize(path))
wb = openpyxl.load_workbook(path, data_only=False)
print('sheets', wb.sheetnames)
for ws in wb.worksheets[:5]:
    print('SHEET', ws.title, 'rows', ws.max_row, 'cols', ws.max_column)
    for row in ws.iter_rows(min_row=1, max_row=min(10, ws.max_row), values_only=True):
        print(row)
