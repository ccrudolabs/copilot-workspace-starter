import openpyxl

wb = openpyxl.load_workbook('GitHub_Copilot_SKUs_Licenciamiento_360_Carlos_Crudo_v7.xlsx', data_only=True)
print('SHEETS', len(wb.sheetnames))
keywords = ('laboratorio', 'taller', 'desarrollador', 'prompt', 'prueba', 'seguridad', 'governanza', 'review', 'agent', 'copilot chat', 'id', 'github')
for name in wb.sheetnames:
    ws = wb[name]
    for row in ws.iter_rows(values_only=True):
        text = ' '.join([str(x) for x in row if x is not None])
        if any(k.lower() in text.lower() for k in keywords):
            print('\nSHEET:', name)
            print('ROW:', row)
            break
